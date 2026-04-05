#!/usr/bin/env python3
"""
Process EU_MedTech_Demo_Complete.xlsx into detailed JSON data files
for the PRYZM dashboard frontend.
"""

import json
import os
import sys
from datetime import datetime, date
from collections import defaultdict

import openpyxl

EXCEL_PATH = "/Users/dharmendersingh/Documents/demo /EU_MedTech_Demo_Complete.xlsx"
OUTPUT_DIR = "/Users/dharmendersingh/Documents/demo/frontend/src/data"


def read_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts using header row."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


def to_iso(val):
    """Convert datetime to ISO date string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return val


def safe_float(val, decimals=2):
    """Round a float safely."""
    if val is None:
        return 0.0
    try:
        return round(float(val), decimals)
    except (ValueError, TypeError):
        return 0.0


def safe_int(val):
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def write_json(filename, data):
    """Write JSON file and return size."""
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w") as f:
        json.dump(data, f, separators=(",", ":"))
    size = os.path.getsize(path)
    return size


def main():
    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    # ─── 1. sales_transactions.json ───
    print("Processing Sales_Transactions...")
    sales_raw = read_sheet(wb, "Sales_Transactions")
    transactions = []
    for r in sales_raw:
        dt = r.get("Transaction Date")
        transactions.append({
            "sku": r["SKU"],
            "description": r["Description"],
            "date": to_iso(dt),
            "year": safe_int(r.get("Year")),
            "month": dt.month if isinstance(dt, (datetime, date)) else 1,
            "unit_price": safe_float(r.get("Unit Price")),
            "quantity": safe_int(r.get("Quantity")),
            "total_value": safe_float(r.get("Total Value")),
            "customer": r.get("Customer Name", ""),
        })
    transactions.sort(key=lambda x: x["date"], reverse=True)
    size = write_json("sales_transactions.json", {"transactions": transactions})
    print(f"  → {len(transactions)} rows, {size / 1024:.1f} KB")

    # ─── 2. products.json ───
    print("Processing Products...")
    categories_raw = read_sheet(wb, "Product_Categories")
    cat_map = {}
    for r in categories_raw:
        cat_map[r["SKU"]] = {
            "category": r.get("Category", ""),
            "sub_category": r.get("Sub_Category", ""),
            "product_type": r.get("Product_Type", ""),
        }

    # Parse gross margin sheets
    margin_data = {}  # sku -> {year: {margin, revenue, units, description}}
    for year, sheet in [(2023, "Gross_Margin_2023"), (2024, "Gross_Margin_2024"), (2025, "Gross_Margin_2025")]:
        rows = read_sheet(wb, sheet)
        for r in rows:
            sku = r["SKU"]
            if sku not in margin_data:
                margin_data[sku] = {}
            margin_data[sku][year] = {
                "margin": safe_float(r.get("Gross Margin - %"), 4),
                "revenue": safe_float(r.get("Total Sales Value")),
                "units": safe_int(r.get("Sales Qty")),
                "description": r.get("Description", ""),
            }

    products = []
    all_skus = set(list(cat_map.keys()) + list(margin_data.keys()))
    for sku in all_skus:
        cat = cat_map.get(sku, {})
        md = margin_data.get(sku, {})

        m23 = md.get(2023, {})
        m24 = md.get(2024, {})
        m25 = md.get(2025, {})

        desc = m25.get("description") or m24.get("description") or m23.get("description") or ""
        margin_2024 = m24.get("margin", 0)
        margin_2025 = m25.get("margin", 0)

        diff = margin_2025 - margin_2024
        if diff > 0.02:
            trend = "up"
        elif diff < -0.02:
            trend = "down"
        else:
            trend = "stable"

        total_rev = safe_float(m23.get("revenue", 0) + m24.get("revenue", 0) + m25.get("revenue", 0))
        total_units = safe_int(m23.get("units", 0) + m24.get("units", 0) + m25.get("units", 0))

        products.append({
            "sku": sku,
            "description": desc,
            "category": cat.get("category", ""),
            "sub_category": cat.get("sub_category", ""),
            "product_type": cat.get("product_type", ""),
            "margin_2023": safe_float(m23.get("margin", 0), 4),
            "margin_2024": safe_float(margin_2024, 4),
            "margin_2025": safe_float(margin_2025, 4),
            "margin_trend": trend,
            "revenue_2023": safe_float(m23.get("revenue", 0)),
            "revenue_2024": safe_float(m24.get("revenue", 0)),
            "revenue_2025": safe_float(m25.get("revenue", 0)),
            "units_2023": safe_int(m23.get("units", 0)),
            "units_2024": safe_int(m24.get("units", 0)),
            "units_2025": safe_int(m25.get("units", 0)),
            "total_revenue": total_rev,
            "total_units": total_units,
            "is_at_risk": margin_2025 is not None and margin_2025 < 0.50,
        })

    products.sort(key=lambda x: x["total_revenue"], reverse=True)
    size = write_json("products.json", {"products": products})
    print(f"  → {len(products)} products, {size / 1024:.1f} KB")

    # ─── 3. customers_detail.json ───
    print("Processing Customer_Master...")
    cust_raw = read_sheet(wb, "Customer_Master")
    ref_date = datetime(2025, 12, 31)
    customers = []
    for r in cust_raw:
        last_purchase = r.get("Last_Purchase_Date")
        days_since = 0
        if isinstance(last_purchase, (datetime, date)):
            days_since = (ref_date - (last_purchase if isinstance(last_purchase, datetime) else datetime.combine(last_purchase, datetime.min.time()))).days

        customers.append({
            "code": r.get("Customer_Name", ""),
            "first_purchase": to_iso(r.get("First_Purchase_Date")),
            "last_purchase": to_iso(last_purchase),
            "lifetime_value": safe_float(r.get("Lifetime_Value")),
            "transaction_count": safe_int(r.get("Transaction_Count")),
            "avg_order_value": safe_float(r.get("Avg_Order_Value")),
            "segment": r.get("Segment", ""),
            "region": r.get("Region", ""),
            "hospital_type": r.get("Hospital_Type", ""),
            "churn_risk": r.get("Churn_Risk", ""),
            "nps": safe_int(r.get("NPS")),
            "days_since_last_purchase": max(days_since, 0),
        })
    customers.sort(key=lambda x: x["lifetime_value"], reverse=True)
    size = write_json("customers_detail.json", {"customers": customers})
    print(f"  → {len(customers)} customers, {size / 1024:.1f} KB")

    # ─── 4. monthly_detail.json ───
    print("Processing Monthly_Margins...")
    monthly_raw = read_sheet(wb, "Monthly_Margins")
    monthly = []
    totals_map = defaultdict(lambda: {
        "total_revenue": 0, "total_cogs": 0, "total_margin_value": 0,
        "margin_pcts": [], "fx_rates": [], "transaction_count": 0, "unique_skus": set()
    })

    for r in monthly_raw:
        year = safe_int(r.get("Year"))
        month = safe_int(r.get("Month"))
        sales_val = safe_float(r.get("Sales_Value"))
        cogs_val = safe_float(r.get("COGS_Value"))
        margin_val = safe_float(r.get("Gross_Margin_Value"))
        margin_pct = safe_float(r.get("Gross_Margin_Pct"), 4)
        fx = safe_float(r.get("FX_Rate"))
        sku = r.get("SKU", "")

        monthly.append({
            "year": year,
            "month": month,
            "sku": sku,
            "description": r.get("Description", ""),
            "sales_qty": safe_float(r.get("Sales_Qty")),
            "sales_value": sales_val,
            "cogs_value": cogs_val,
            "margin_value": margin_val,
            "margin_pct": margin_pct,
            "fx_rate": fx,
        })

        key = (year, month)
        t = totals_map[key]
        t["total_revenue"] += sales_val
        t["total_cogs"] += cogs_val
        t["total_margin_value"] += margin_val
        if margin_pct > 0:
            t["margin_pcts"].append(margin_pct)
        if fx > 0:
            t["fx_rates"].append(fx)
        t["transaction_count"] += 1
        t["unique_skus"].add(sku)

    monthly_totals = []
    for (year, month), t in sorted(totals_map.items()):
        avg_margin = sum(t["margin_pcts"]) / len(t["margin_pcts"]) if t["margin_pcts"] else 0
        avg_fx = sum(t["fx_rates"]) / len(t["fx_rates"]) if t["fx_rates"] else 0
        monthly_totals.append({
            "year": year,
            "month": month,
            "total_revenue": safe_float(t["total_revenue"]),
            "total_cogs": safe_float(t["total_cogs"]),
            "total_margin_value": safe_float(t["total_margin_value"]),
            "avg_margin_pct": safe_float(avg_margin, 4),
            "avg_fx_rate": safe_float(avg_fx),
            "transaction_count": t["transaction_count"],
            "unique_skus": len(t["unique_skus"]),
        })

    size = write_json("monthly_detail.json", {"monthly": monthly, "monthly_totals": monthly_totals})
    print(f"  → {len(monthly)} monthly rows, {len(monthly_totals)} monthly totals, {size / 1024:.1f} KB")

    # ─── 5. cogs_detail.json ───
    print("Processing COGS...")
    cogs_raw = read_sheet(wb, "COGS")
    cogs = []
    for r in cogs_raw:
        cogs.append({
            "sku": r.get("SKU", ""),
            "description": r.get("Description", ""),
            "qty": safe_int(r.get("Qty.")),
            "unit_price_eur": safe_float(r.get("Unit price (EUR)")),
            "total_price_eur": safe_float(r.get("Total Price (EUR)")),
            "fx_rate": safe_float(r.get("FX Rate (EUR/INR)")),
            "landed_unit_cost": safe_float(r.get("Landed Unit Cost (EUR)")),
            "landed_total": safe_float(r.get("Landed Total (EUR)")),
            "freight_eur": safe_float(r.get("Freight (EUR)")),
            "duty_eur": safe_float(r.get("Duty (EUR)")),
            "total_eur": safe_float(r.get("Total (EUR)")),
            "year": safe_int(r.get("Year")),
        })
    size = write_json("cogs_detail.json", {"cogs": cogs})
    print(f"  → {len(cogs)} COGS rows, {size / 1024:.1f} KB")

    # ─── 6. pipeline.json ───
    print("Processing CRM_Pipeline...")
    pipeline_raw = read_sheet(wb, "CRM_Pipeline")
    pipeline = []
    stage_agg = defaultdict(lambda: {"count": 0, "total_value": 0, "cycles": []})
    rep_agg = defaultdict(lambda: {"won_count": 0, "won_value": 0, "lost_count": 0, "pipeline_count": 0})

    total_won = 0
    total_closed = 0

    for r in pipeline_raw:
        stage = r.get("Stage", "")
        deal_val = safe_float(r.get("Deal_Value"))
        cycle = safe_int(r.get("Sales_Cycle_Days"))
        rep = r.get("Sales_Rep", "")

        pipeline.append({
            "deal_id": r.get("Deal_ID", ""),
            "customer": r.get("Customer_Name", ""),
            "sku": r.get("SKU", ""),
            "description": r.get("Description", ""),
            "sales_rep": rep,
            "stage": stage,
            "win_probability": safe_float(r.get("Win_Probability"), 2),
            "deal_value": deal_val,
            "start_date": to_iso(r.get("Start_Date")),
            "expected_close": to_iso(r.get("Expected_Close_Date")),
            "sales_cycle_days": cycle,
        })

        sa = stage_agg[stage]
        sa["count"] += 1
        sa["total_value"] += deal_val
        if cycle > 0:
            sa["cycles"].append(cycle)

        if stage == "Closed Won":
            rep_agg[rep]["won_count"] += 1
            rep_agg[rep]["won_value"] += deal_val
            total_won += 1
            total_closed += 1
        elif stage == "Closed Lost":
            rep_agg[rep]["lost_count"] += 1
            total_closed += 1
        else:
            rep_agg[rep]["pipeline_count"] += 1

    by_stage = []
    for stage, sa in stage_agg.items():
        by_stage.append({
            "stage": stage,
            "count": sa["count"],
            "total_value": safe_float(sa["total_value"]),
            "avg_cycle": safe_float(sum(sa["cycles"]) / len(sa["cycles"])) if sa["cycles"] else 0,
        })

    by_rep = []
    for rep, ra in rep_agg.items():
        by_rep.append({
            "rep": rep,
            "won_count": ra["won_count"],
            "won_value": safe_float(ra["won_value"]),
            "lost_count": ra["lost_count"],
            "pipeline_count": ra["pipeline_count"],
        })
    by_rep.sort(key=lambda x: x["won_value"], reverse=True)

    all_values = [p["deal_value"] for p in pipeline]
    all_cycles = [p["sales_cycle_days"] for p in pipeline if p["sales_cycle_days"] > 0]

    pipeline_summary = {
        "by_stage": by_stage,
        "by_rep": by_rep,
        "win_rate": safe_float(total_won / total_closed, 4) if total_closed > 0 else 0,
        "avg_deal_value": safe_float(sum(all_values) / len(all_values)) if all_values else 0,
        "avg_cycle_days": safe_float(sum(all_cycles) / len(all_cycles)) if all_cycles else 0,
    }

    size = write_json("pipeline.json", {"pipeline": pipeline, "pipeline_summary": pipeline_summary})
    print(f"  → {len(pipeline)} deals, {len(by_rep)} reps, {size / 1024:.1f} KB")

    # ─── 7. inventory_detail.json ───
    print("Processing Inventory_Snapshot...")
    inv_raw = read_sheet(wb, "Inventory_Snapshot")
    inventory = []
    for r in inv_raw:
        inventory.append({
            "sku": r.get("SKU", ""),
            "description": r.get("Description", ""),
            "category": r.get("Category", ""),
            "product_type": r.get("Product_Type", ""),
            "avg_monthly_demand": safe_float(r.get("Avg_Monthly_Demand")),
            "lead_time": safe_int(r.get("Lead_Time_Days")),
            "safety_stock": safe_int(r.get("Safety_Stock")),
            "reorder_point": safe_int(r.get("Reorder_Point")),
            "current_stock": safe_int(r.get("Current_Stock")),
            "status": r.get("Status", ""),
            "unit_cost_eur": safe_float(r.get("Unit_Cost_EUR")),
        })
    size = write_json("inventory_detail.json", {"inventory": inventory})
    print(f"  → {len(inventory)} inventory items, {size / 1024:.1f} KB")

    # ─── 8. price_governance.json ───
    print("Processing Price_Governance...")
    gov_raw = read_sheet(wb, "Price_Governance")
    governance = []
    for r in gov_raw:
        governance.append({
            "category": r.get("Category", ""),
            "sub_category": r.get("Sub_Category", ""),
            "product_type": r.get("Product_Type", ""),
            "sku_count": safe_int(r.get("SKU_Count")),
            "target_margin": safe_float(r.get("Target_Margin_Pct"), 4),
            "max_discount": safe_float(r.get("Max_Discount_Pct"), 4),
            "auto_approve": safe_float(r.get("Auto_Approve_Threshold"), 4),
            "manager_approve": safe_float(r.get("Manager_Approve_Threshold"), 4),
            "director_approve": safe_float(r.get("Director_Approve_Threshold"), 4),
            "vp_approve": safe_float(r.get("VP_Approve_Above"), 4),
            "review_frequency": r.get("Review_Frequency", ""),
        })
    size = write_json("price_governance.json", {"governance": governance})
    print(f"  → {len(governance)} governance rules, {size / 1024:.1f} KB")

    # ─── 9. forecasting.json ───
    print("Generating Forecasting data...")

    # Get 2025 monthly totals
    actuals_2025 = sorted(
        [t for t in monthly_totals if t["year"] == 2025],
        key=lambda x: x["month"]
    )

    # Build forecast vs actuals for 2025
    # Use 3-month rolling avg as "model prediction"
    forecast_vs_actuals = []
    for i, mt in enumerate(actuals_2025):
        actual = mt["total_revenue"]

        # 3-month rolling average of prior months as forecast
        if i >= 3:
            window = [actuals_2025[j]["total_revenue"] for j in range(i - 3, i)]
            forecast = sum(window) / 3
        elif i >= 1:
            window = [actuals_2025[j]["total_revenue"] for j in range(0, i)]
            forecast = sum(window) / len(window)
        else:
            # First month: use 2024 Dec or just use actual as baseline
            dec_2024 = [t for t in monthly_totals if t["year"] == 2024 and t["month"] == 12]
            forecast = dec_2024[0]["total_revenue"] if dec_2024 else actual

        is_shock = mt["month"] in [5, 10]

        # For shock months, the model predicts normal levels
        if is_shock:
            # Use surrounding months to estimate what "normal" would have been
            nearby = [actuals_2025[j]["total_revenue"] for j in range(max(0, i-2), min(len(actuals_2025), i+3)) if j != i]
            forecast = sum(nearby) / len(nearby) if nearby else forecast

        forecast_vs_actuals.append({
            "month": mt["month"],
            "actual_revenue": safe_float(actual),
            "forecast_revenue": safe_float(forecast),
            "p10": safe_float(forecast * 0.6),
            "p90": safe_float(forecast * 1.4),
            "is_shock": is_shock,
        })

    # H1 2026 forecast
    h1_2026 = []
    growth_factors = {1: 1.15, 2: 1.17, 3: 1.25, 4: 1.18, 5: 1.16, 6: 1.19}
    for m in range(1, 7):
        # Find 2025 same-month actual
        same_month = [t for t in actuals_2025 if t["month"] == m]
        if same_month:
            baseline = same_month[0]["total_revenue"]
            # For shock months (5), use average of surrounding months instead
            if m == 5:
                nearby = [t["total_revenue"] for t in actuals_2025 if t["month"] in [3, 4, 6, 7]]
                baseline = sum(nearby) / len(nearby) if nearby else baseline
        else:
            baseline = sum(t["total_revenue"] for t in actuals_2025) / len(actuals_2025) if actuals_2025 else 200000

        factor = growth_factors.get(m, 1.17)
        p50 = baseline * factor

        h1_2026.append({
            "month": m,
            "p10": safe_float(p50 * 0.75),
            "p50": safe_float(p50),
            "p90": safe_float(p50 * 1.30),
        })

    size = write_json("forecasting.json", {
        "forecast_vs_actuals_2025": forecast_vs_actuals,
        "h1_2026_forecast": h1_2026,
    })
    print(f"  → {len(forecast_vs_actuals)} months forecast vs actuals, {len(h1_2026)} months H1 2026, {size / 1024:.1f} KB")

    wb.close()

    # ─── Summary ───
    print("\n" + "=" * 50)
    print("All JSON files generated in src/data/:")
    total_size = 0
    for f in sorted(os.listdir(OUTPUT_DIR)):
        if f.endswith(".json"):
            fsize = os.path.getsize(os.path.join(OUTPUT_DIR, f))
            total_size += fsize
            print(f"  {f:40s} {fsize / 1024:8.1f} KB")
    print(f"  {'TOTAL':40s} {total_size / 1024:8.1f} KB")
    print("=" * 50)


if __name__ == "__main__":
    main()
